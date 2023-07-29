from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QPixmap, QImage
import webbrowser
import xml.etree.ElementTree as xml
import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
from scripts.popup import Ui_Form
import threading
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from main_interface import stop_value

class Face_infos():  # leave this empty
    def __init__(self):  # constructor function using self
        self.Name = None  # variable using self.
        self.Encodings = None  # variable using self


# had to Disable the window separately
def multi_threads(passedArg):
    def openPopup():
        passedArg.window = QtWidgets.QMainWindow()
        passedArg.ui = Ui_Form()
        passedArg.ui.setupUi(passedArg.window)
        passedArg.window.show()
        return passedArg.window
    new_Window=openPopup()
    passedArg.centralwidget.setEnabled(False)
    thread1 = threading.Thread(target=generateXML, args=("encodings.xml",new_Window,passedArg))
    thread1.start()

# encode images
def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encoding = face_recognition.face_encodings(img)[0]
        encodeList.append(encoding)
    return encodeList

# Function that Creates XML file of Encodings
def generateXML(filename,new_Window,passedSelf):

    root = xml.Element("Faces")
    soldiersImages = []
    soldiersNames = []
    myData = os.listdir('imagesDB')
    print(myData)

    size = len(myData)
    # iterate over all the images in the folder
    for data in myData:
        curImage = cv2.imread(f'imagesDB/{data}')
        if curImage is not None:
            soldiersImages.append(curImage)
            soldiersNames.append(data.split('.')[0])
        else:
            print(f"Error: Unable to load the image '{data}'")

        # soldiersImages.append(curImage)
        # soldiersNames.append(data.split('.')[0])
    
    # save encondings in array
    encodedListKnown = findEncodings(soldiersImages)
    print('Encoded Successfully')

    # Create XML structure
    for index, name in enumerate(soldiersNames):
        face = xml.SubElement(root, "face")
        name = xml.SubElement(face, "name")
        name.text = soldiersNames[index]
        encodings = xml.SubElement(face, "encodings")

        print(f'{index+1} out of {size} images.')
        for encode in encodedListKnown[index]:
            encoding = xml.SubElement(encodings, "encoding")
            encoding.text = str(encode)

    # save xml struct as a file
    tree = xml.ElementTree(root)
    with open(filename, 'wb') as files:
        tree.write(files)
        print("Refreshed Successfully")
    new_Window.close()
    passedSelf.centralwidget.setEnabled(True)

# Starting the APP
def readtree(startBtn,stopBtn,refreshBtn,display_Label,spinBox):
    global stop_value
    stop_value=0
    startBtn.setEnabled(False)
    stopBtn.setEnabled(True)
    refreshBtn.setEnabled(False)
    # iterate my XML
    tree = xml.parse("encodings.xml")
    root = tree.getroot()
    faces = []
    for face in root.findall('face'):
        face_infos = Face_infos()
        face_encodings = []
        for name_node in face.findall('name'):
            face_infos.Name = name_node.text
        for encodings in face.findall('encodings'):
            for encoding in encodings.findall('encoding'):
                face_encodings.append(float(encoding.text))
            face_infos.Encodings = face_encodings
        faces.append(face_infos)
    encodedListKnown = []
    soldiersNames = []
    for element in faces:
        print(element.Name)
        soldiersNames.append(element.Name)
        encodedListKnown.append(element.Encodings)

    camera = spinBox.value()
    cap = cv2.VideoCapture(int(camera))

    while stop_value==0:
        success, img = cap.read()
        imgScaled = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgScaled = cv2.cvtColor(imgScaled, cv2.COLOR_BGR2RGB)

        facesCurrentFrame = face_recognition.face_locations(imgScaled)
        encodingCurrentFrame = face_recognition.face_encodings(imgScaled, facesCurrentFrame)

        for encodedFace, faceLoc in zip(encodingCurrentFrame, facesCurrentFrame):
            matches = face_recognition.compare_faces(encodedListKnown, encodedFace)
            faceDist = face_recognition.face_distance(encodedListKnown, encodedFace)
            matchIndex = np.argmin(faceDist)

            if matches[matchIndex]:
                success, img = cap.read()

                name = soldiersNames[matchIndex].upper()
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 255), 1)

                # Top Left  x,y
                cv2.line(img, (x1, y1), (x1 + 30, y1), (0, 255, 0), 5)
                cv2.line(img, (x1, y1), (x1, y1 + 30), (0, 255, 0), 5)
                # Top Right  x1,y
                cv2.line(img, (x2, y1), (x2 - 30, y1), (0, 255, 0), 5)
                cv2.line(img, (x2, y1), (x2, y1 + 30), (0, 255, 0), 5)
                # Bottom Left  x,y1
                cv2.line(img, (x1, y2), (x1 + 30, y2), (0, 255, 0), 5)
                cv2.line(img, (x1, y2), (x1, y2 - 30), (0, 255, 0), 5)
                # Bottom Right  x1,y1
                cv2.line(img, (x2, y2), (x2 - 30, y2), (0, 255, 0), 5)
                cv2.line(img, (x2, y2), (x2, y2 - 30), (0, 255, 0), 5)

                cv2.putText(img, name, (x1 + 6, y2 + 25), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 100, 0), 1,cv2.LINE_4)

                markAttendance(name)
            else:

                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 0, 255), 1)

                # Top Left  x,y
                cv2.line(img, (x1, y1), (x1 + 30, y1), (0, 255, 255), 5)
                cv2.line(img, (x1, y1), (x1, y1 + 30), (0, 255, 255), 5)
                # Top Right  x1,y
                cv2.line(img, (x2, y1), (x2 - 30, y1), (0, 255, 255), 5)
                cv2.line(img, (x2, y1), (x2, y1 + 30), (0, 255, 255), 5)
                # Bottom Left  x,y1
                cv2.line(img, (x1, y2), (x1 + 30, y2), (0, 255, 255), 5)
                cv2.line(img, (x1, y2), (x1, y2 - 30), (0, 255, 255), 5)
                # Bottom Right  x1,y1
                cv2.line(img, (x2, y2), (x2 - 30, y2), (0, 255, 255), 5)
                cv2.line(img, (x2, y2), (x2, y2 - 30), (0, 255, 255), 5)

                cv2.putText(img, "Unknown", (x1 + 6, y2 + 25), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 150), 1,cv2.LINE_4)

        display_image(img,display_Label,1)
        cv2.waitKey()
    cap.release()
    display_Label.setPixmap(QtGui.QPixmap("./images/wave_background.svg"))



# Stop Function
def stop(startBtn,stopBtn,refreshBtn):
    startBtn.setEnabled(True)
    stopBtn.setEnabled(False)
    refreshBtn.setEnabled(True)
    global stop_value
    stop_value=1

# Display in the label instead of cv2.imShow

def display_image(img,display_Label,window=1):
    qformat = QImage.Format_Indexed8
    if len(img.shape) == 3:
        if img.shape[2] == 4:
            qformat = QImage.Format_RGBA8888
        else:
            qformat = QImage.Format_RGB888
    outImage = QImage(img, img.shape[1], img.shape[0], img.strides[0], qformat)
    outImage = outImage.rgbSwapped()

    display_Label.setPixmap(QPixmap.fromImage(outImage))

# HyperLink function
def doSomething(click):
    webbrowser.open('http://instagram.com/yacinezerimi', new=2)

# Attendance
def markAttendance(name):

    today_date=datetime.today()
    today_date= today_date.strftime("%B %d, %Y")

    if os.path.isfile(f'attendance/{today_date}.xlsx'):
        print('file Exists')
        work_book = openpyxl.load_workbook(f'attendance/{today_date}.xlsx')
        work_sheet = work_book.active
    else:
        work_book = openpyxl.Workbook()
        work_sheet =  work_book.active
        work_sheet.title = "Main"
        c1 = work_sheet.cell(row= 1 , column = 1)
        c1.font = Font(size = 18 )
        c1.value = f"Attendance for Day : {today_date}"
        c1 = work_sheet.cell(row= 2 , column = 1)
        c1.value = "Full Name"
        c1.fill = PatternFill(bgColor="8cc963", fill_type = "solid")
        c1.alignment = Alignment(horizontal='center')
        c1.font = Font(bold=True)
        c1 = work_sheet.cell(row= 2 , column = 2)
        c1.value = "Time Arrived"
        c1.fill = PatternFill(bgColor="8cc963", fill_type = "solid")
        c1.alignment = Alignment(horizontal='center')
        c1.font = Font(bold=True)
        work_sheet.merge_cells('A1:D1')
        work_sheet.row_dimensions[1].height = 30
        work_sheet.row_dimensions[2].height = 22
        work_sheet.column_dimensions['A'].width = 20
        work_sheet.column_dimensions['B'].width = 20

        work_book.save(f'attendance/{today_date}.xlsx')

        print('file DOES NOT Exists, Created now.')

    # get attendance list
    names_attended=[]
    for i in range(2,work_sheet.max_row+1):
        cell_obj = work_sheet.cell(row = i, column = 1)
        names_attended.append(cell_obj.value)

    # check attendance and append new attendance
    if name not in names_attended:
        cell_obj = work_sheet.cell(row = work_sheet.max_row+1, column = 1)
        cell_obj.value = name
        cell_obj = work_sheet.cell(row = work_sheet.max_row, column = 2)
        now = datetime.now()
        dtString = now.strftime('%H:%M:%S')
        cell_obj.value = dtString
    # formatting cells
    if work_sheet.max_row > 3:
        for l in range(1,work_sheet.max_row):
            if (l % 2) == 0:
                cell_obj = work_sheet.cell(row = l, column = 1)
                cell_obj.fill = PatternFill(bgColor="8cc963", fill_type = "solid")
                cell_obj = work_sheet.cell(row = l, column = 2)
                cell_obj.fill = PatternFill(bgColor="8cc963", fill_type = "solid")
    work_book.save(f'attendance/{today_date}.xlsx')

# open today's attendance
def open_Attendance():
    today_date=datetime.today()
    today_date= today_date.strftime("%B %d, %Y")
    file_name = today_date.split()
    new_name = file_name[0]+file_name[1]+file_name[2]+'.xlsx'
    os.startfile((f"attendance\{file_name[0]} {file_name[1]} {file_name[2]}.xlsx"),'open')
